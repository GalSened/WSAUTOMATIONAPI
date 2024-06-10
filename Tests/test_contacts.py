import base64
import random
import unittest
import uuid
import warnings
from pathlib import Path
from pprint import pprint
from time import sleep
import openpyxl
import pytest
import json
from shared import Shared
from Enums.status_codes import StatusCode, ResultCode
from Common.all_api_methods import WesignMethodsApi

@pytest.mark.flaky(max_runs=3)
class WesignContactsApi(unittest.TestCase):
    def setUp(self):
        # p = Path(__file__).with_name('ContactsSettings.json')
        # with open(p) as f:
        #     self.settings = json.load(f)
        p = Path(__file__).resolve().parent.parent
        file_path = p / "Settings\\ContactsSettings.json"
        with open(file_path) as f:
            self.settings = json.load(f)
        warnings.simplefilter('ignore', ResourceWarning)
        warnings.simplefilter('ignore', DeprecationWarning)
        self.token = Shared.login_request(self)
        ##test

    @pytest.mark.part1
    def test_create_new_contact_by_email_success(self):
        email_prefix = uuid.uuid4().hex
        self.email = email_prefix + "@comda.co.il"
        random_hex = uuid.uuid4().hex
        phone_number = ''.join(filter(str.isdigit, random_hex))[:10]
        with open(self.settings["CreateNewValidContactWithEmailAndPhone"], 'r+') as f:
            data = json.load(f)
            data["email"] = self.email  # <--- add `id` value.
            data["phone"] = phone_number
            f.seek(0)  # <--- should reset file position to the beginning.
            json.dump(data, f, indent=3)
            f.truncate()  # remove remaining part
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewValidContactWithEmailAndPhone')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactId']
        assert len(json_response) > 0
        r = WesignMethodsApi.contacts_id_delete(self, json_response)
        assert r.status_code == StatusCode.OK

    @pytest.mark.part2
    def test_create_new_contact_by_phone_success(self):
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewValidContactSendingBySms')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactId']
        assert len(json_response) > 0
        r = WesignMethodsApi.contacts_id_delete(self, json_response)
        assert r.status_code == StatusCode.OK

    @pytest.mark.part3
    def test_create_new_contact_by_invalid_email(self):
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewContactWithInvalidEmail')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['DefaultSendingMethod']
        assert json_response[0] == ResultCode.DEFAULT_SENDING_METHOD

    @pytest.mark.part1
    def test_create_new_contact_by_invalid_phone(self):
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewContactWithInvalidPhone')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['Phone']
        assert json_response[0] == ResultCode.PLEASE_SPECIFY_VALID_PHONE

    @pytest.mark.part2
    def test_create_new_contact_by_empty_name(self):
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewContactWithEmptyName')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['Name']
        assert json_response == ResultCode.INVALID_NAME

    @pytest.mark.part3
    def test_create_new_contact_by_phone_invalid_sending_type(self):
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewValidContactByPhoneInvalidSendingMethod')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['DefaultSendingMethod']
        assert json_response[0] == ResultCode.DEFAULT_SENDING_METHOD

    @pytest.mark.part1
    def test_create_new_contact_by_email_invalid_sending_type(self):
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewValidContactByEmailInvalidSendingMethod')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['DefaultSendingMethod']
        assert json_response[0] == ResultCode.DEFAULT_SENDING_METHOD

    @pytest.mark.part2
    def test_create_new_contact_by_email_invalid_default_sending_method(self):
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewValidContactByEmailInvalidSendingMethod')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['DefaultSendingMethod']
        assert json_response[0] == ResultCode.DEFAULT_SENDING_METHOD

    @pytest.mark.part3
    def test_create_new_contact_with_name_already_exists(self):
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewValidContactNameExists')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['error']
        assert json_response[0] == ResultCode.CONTACT_WITH_SAME_MEANS_ALREADY_EXISTS

    @pytest.mark.part1
    def test_create_new_contact_with_email_already_exists(self):
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewValidContactEmailExists')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['error']
        assert json_response[0] == ResultCode.CONTACT_WITH_SAME_MEANS_ALREADY_EXISTS

    @pytest.mark.part2
    def test_create_new_contact_with_phone_already_exists(self):
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewValidContactPhoneExists')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['error']
        assert json_response[0] == ResultCode.CONTACT_WITH_SAME_MEANS_ALREADY_EXISTS

    @pytest.mark.part3
    def test_create_new_contact_from_bulk_xlsx_file_success(self):
        r = WesignMethodsApi.contacts_bulk_post_json_file(self, 'CreateNewContactFromBulkCsv')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactsId']
        assert len(json_response[0]) == 36
        r = WesignMethodsApi.contacts_id_delete(self, json_response[0])
        assert r.status_code == StatusCode.OK

    @pytest.mark.part1
    def test_create_new_contacts_from_bulk_xlsx_sending_method_sms_success(self):
        r = WesignMethodsApi.contacts_bulk_post_json_file(self, 'CreateNewContactsFromBulkCsvSendingMehtodSms')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactsId']
        assert len(json_response[0]) == 36
        for contacts in json_response:
            WesignMethodsApi.contacts_id_delete(self, contacts)

    @pytest.mark.part2
    def test_create_new_contacts_from_bulk_xlsx_sending_method_email_success(self):
        r = WesignMethodsApi.contacts_bulk_post_json_file(self, 'CreateNewContactsFromBulkCsvSendingMehtodEmail')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactsId']
        assert len(json_response[0]) == 36
        for contacts in json_response:
            WesignMethodsApi.contacts_id_delete(self, contacts)

    @pytest.mark.part3
    def test_create_new_contacts_from_bulk_xlsx_sending_method_email_and_sms_success(self):
        r = WesignMethodsApi.contacts_bulk_post_json_file(self, 'CreateNewContactsFromBulkCsvSendingMehtodEmailAndSms')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactsId']
        assert len(json_response[0]) == 36
        for contacts in json_response:
            WesignMethodsApi.contacts_id_delete(self, contacts)

    @pytest.mark.part1
    def test_create_new_contact_from_bulk_xlsx_valid_name_and_email_invalid_phone(self):
        r = WesignMethodsApi.contacts_bulk_post_json_file(self, 'CreateNewContactsFromBulkCsvInvalidNameAndEmailInvalidPhone')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['error']
        assert json_response[0] == ResultCode.INVALID_PHONE

    @pytest.mark.part2
    def test_create_new_contact_from_bulk_xlsx_empty_csv(self):
        r = WesignMethodsApi.contacts_bulk_post_json_file(self, 'CreateNewContactsFromBulkCsvEmpty')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['error']
        assert json_response[0] == ResultCode.INVALID_FORMAT

    @pytest.mark.part3
    def test_create_new_contact_from_bulk_csv_empty_fullname_valid_phone(self):
        r = WesignMethodsApi.contacts_bulk_post_json_file(self, 'CreateNewContactsFromBulkCsvEmptyFullnameValidPhone')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['error']
        assert json_response[0] == ResultCode.NAME_IS_MISSING

    @pytest.mark.part1
    def test_update_contact_mail(self):
        email_prefix = uuid.uuid4().hex
        self.email = email_prefix + "@comda.co.il"
        random_hex = uuid.uuid4().hex
        phone_number = ''.join(filter(str.isdigit, random_hex))[:10]
        with open(self.settings["CreateNewValidContactWithEmailAndPhone"], 'r+') as f:
            data = json.load(f)
            data["email"] = self.email  # <--- add `id` value.
            data["phone"] = phone_number
            f.seek(0)  # <--- should reset file position to the beginning.
            json.dump(data, f, indent=3)
            f.truncate()  # remove remaining part
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewValidContactWithEmailAndPhone')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactId']
        r = WesignMethodsApi.contacts_id_put(self, json_response, 'UpdateContactMail')
        assert r.status_code == StatusCode.OK
        WesignMethodsApi.contacts_id_delete(self, json_response)
        assert r.status_code == StatusCode.OK

    @pytest.mark.part2
    def test_update_contact_phone(self):
        email_prefix = uuid.uuid4().hex
        self.email = email_prefix + "@comda.co.il"
        random_hex = uuid.uuid4().hex
        phone_number = ''.join(filter(str.isdigit, random_hex))[:10]
        with open(self.settings["CreateNewValidContactWithEmailAndPhone"], 'r+') as f:
            data = json.load(f)
            data["email"] = self.email  # <--- add `id` value.
            data["phone"] = phone_number
            f.seek(0)  # <--- should reset file position to the beginning.
            json.dump(data, f, indent=3)
            f.truncate()  # remove remaining part
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewValidContactWithEmailAndPhone')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactId']
        r = WesignMethodsApi.contacts_id_put(self, json_response, 'UpdateContactPhone')
        assert r.status_code == StatusCode.OK
        WesignMethodsApi.contacts_id_delete(self, json_response)
        assert r.status_code == StatusCode.OK

    @pytest.mark.part3
    def test_update_contact_seal(self):
        email_prefix = uuid.uuid4().hex
        self.email = email_prefix + "@comda.co.il"
        random_hex = uuid.uuid4().hex
        phone_number = ''.join(filter(str.isdigit, random_hex))[:10]
        with open(self.settings["CreateNewValidContactWithEmailAndPhone"], 'r+') as f:
            data = json.load(f)
            data["email"] = self.email  # <--- add `id` value.
            data["phone"] = phone_number
            f.seek(0)  # <--- should reset file position to the beginning.
            json.dump(data, f, indent=3)
            f.truncate()  # remove remaining part
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewValidContactWithEmailAndPhone')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactId']
        r = WesignMethodsApi.contacts_id_put(self, json_response, 'UpdateContactSeal')
        assert r.status_code == StatusCode.OK
        WesignMethodsApi.contacts_id_delete(self, json_response)
        assert r.status_code == StatusCode.OK

    @pytest.mark.part1
    def test_update_contact_invalid_email(self):
        email_prefix = uuid.uuid4().hex
        self.email = email_prefix + "@comda.co.il"
        random_hex = uuid.uuid4().hex
        phone_number = ''.join(filter(str.isdigit, random_hex))[:10]
        with open(self.settings["CreateNewValidContactWithEmailAndPhone"], 'r+') as f:
            data = json.load(f)
            data["email"] = self.email  # <--- add `id` value.
            data["phone"] = phone_number
            f.seek(0)  # <--- should reset file position to the beginning.
            json.dump(data, f, indent=3)
            f.truncate()  # remove remaining part
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewValidContactWithEmailAndPhone')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactId']
        r = WesignMethodsApi.contacts_id_put(self, json_response, 'UpdateContactInvalidPhone')
        assert r.status_code == StatusCode.BAD_REQUEST
        response_error = r.json()
        json_response_error = response_error['errors']['Phone']
        assert json_response_error[0] == ResultCode.PLEASE_SPECIFY_VALID_PHONE
        r = WesignMethodsApi.contacts_id_delete(self, json_response)
        assert r.status_code == StatusCode.OK

    @pytest.mark.skip('Not implement in 3.0.8 version')
    def test_update_contact_invalid_name(self):
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewValidContactWithEmailAndPhone')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactId']
        r = WesignMethodsApi.contacts_id_put(self, json_response, 'UpdateContactInvalidName')
        assert r.status_code == StatusCode.BAD_REQUEST
        response_error = r.json()
        json_response_error = response_error['errors']['Name']
        assert json_response_error[0] == ResultCode.NAME_SHOULD_CONTAIN_ONLY_CHARACTERS
        r = WesignMethodsApi.contacts_id_delete(self, json_response)
        assert r.status_code == StatusCode.OK

    #Bug number = WES-1047
    @pytest.mark.part2
    def test_create_new_contact_with_global_number_when_provider_is_goldman(self):
        r = WesignMethodsApi.contacts_post_json_file(self, 'CreateNewContactWithGlobalPhoneWhenProviderIsGoldman')
        assert r.status_code == StatusCode.BAD_REQUEST
        response_error = r.json()
        json_response_error = response_error['errors']['error']
        assert json_response_error[0] == ResultCode.SMS_PROVIDER_ERROR

    @pytest.mark.part3
    def test_delete_multi_contacts_by_id_success(self):
        a = []
        for x in range(4):
            contact = uuid.uuid4().hex
            self.contact_name = contact
            d = {
                  "name": self.contact_name,
                  "email": self.contact_name + "@ApiUser.com",
                  "defaultSendingMethod": 2
                }
            r = WesignMethodsApi.contacts_post(self, d)
            assert r.status_code == StatusCode.OK
            response = r.json()
            json_response = response['contactId']
            assert len(json_response) > 0
            a.append(response['contactId'])
        c = {
              "ids": [
                    a[0], a[1], a[2], a[3]
              ]
            }
        delete = WesignMethodsApi.contacts_delete_batch_put(self, c)
        assert delete.status_code == StatusCode.OK

    @pytest.mark.part1
    def test_upload_hebrew_contacts_from_xlsx_file_and_check_names(self):
        number_of_contacts = 10
        list_of_names = self.__api_get_list_of_random_hebrew_names(number_of_contacts)
        list_of_emails = self.__api_get_list_of_random_emails(number_of_contacts)
        book = openpyxl.load_workbook(self.settings['openXlsxFileForUploadContacts'])
        sheet = book.active
        for i in range(number_of_contacts):
            sheet.cell(row=i+2, column=1).value = list_of_names[i]  # FullName
            sheet.cell(row=i + 2, column=2).value = list_of_emails[i]  # Email
            sheet.cell(row=i + 2, column=4).value = 2
        book.save(self.settings['openXlsxFileForUploadContacts'])
        data = open(self.settings['openXlsxFileForUploadContacts'], 'rb').read()
        base64_encoded = base64.b64encode(data).decode('UTF-8')
        base64_string = '"base64File": "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,' + base64_encoded
        contacts_id = WesignMethodsApi.contacts_bulk_post_xlsx_file(self, base64_string)
        assert contacts_id.status_code == StatusCode.OK
        response = contacts_id.json()
        contacts_list = response['contactsId']
        self.__check_hebrew_contacts(contacts_list, number_of_contacts)
        sheet.delete_rows(2, number_of_contacts)  # Delete contacts from xlsx file
        book.save(self.settings['openXlsxFileForUploadContacts'])
        for contact_id in contacts_list:
            WesignMethodsApi.contacts_id_delete(self, contact_id)

    ##WES-1458
    @pytest.mark.part2
    def test_create_new_contact_group(self, delete=True):
        contact_group = WesignMethodsApi.contacts_group_post(self, 'CreateNewContactGroup')
        assert contact_group.status_code == StatusCode.OK
        response = contact_group.json()
        id = response['id']
        if delete:
            delete_contact_group = WesignMethodsApi.contacts_group_delete(self, id)
            assert delete_contact_group.status_code == StatusCode.OK
        else:
            pass
        return response['id']

    @pytest.mark.part3
    def test_create_new_contact_empty_name_group(self):
        contact_group = WesignMethodsApi.contacts_group_post(self, 'CreateNewContactGroupEmptyName')
        assert contact_group.status_code == StatusCode.BAD_REQUEST
        response = contact_group.json()
        json_response = response['errors']['error']
        assert json_response[0] == ResultCode.INVALID_CONTACTS_GROUP_NAME

    @pytest.mark.part1
    def test_edit_contact_group(self):
        group = uuid.uuid4().hex
        self.group_name = group
        group_id = self.test_create_new_contact_group(False)
        request = {
              "name": self.group_name,
              "contactsGroupMembers": [
                {
                  "contactId": "9e5268d6-1fdb-477c-afef-08db688925a9",
                  "order": 1
                }
              ]
            }
        edit_contact_group = WesignMethodsApi.contacts_group_put(self, request, group_id)
        assert edit_contact_group.status_code == StatusCode.OK
        WesignMethodsApi.contacts_group_delete(self,group_id)

    @pytest.mark.part2
    def test_delete_contact_group(self):
        group = uuid.uuid4().hex
        self.group_name = group
        group_id = self.test_create_new_contact_group(False)
        delete_contact_group = WesignMethodsApi.contacts_group_delete(self, group_id)
        assert delete_contact_group.status_code == StatusCode.OK

    @pytest.mark.part3
    def test_get_contact_group_by_id(self):
        group = uuid.uuid4().hex
        self.group_name = group
        group_id = self.test_create_new_contact_group(False)
        get_contact_group = WesignMethodsApi.contacts_group_get(self, group_id)
        assert get_contact_group.status_code == StatusCode.OK
        response = get_contact_group.json()
        list_of_contacts_id = []
        index = 0
        for x in range(7):
            contacts_ids = response['contactsGroupMembers'][index]['id']
            list_of_contacts_id.append(contacts_ids)
            index += 1

        assert len(list_of_contacts_id) == 7
        delete_contact_group = WesignMethodsApi.contacts_group_delete(self, group_id)
        assert delete_contact_group.status_code == StatusCode.OK

    @pytest.mark.part1
    def test_get_all_contact_group(self):
        groups = WesignMethodsApi.contacts_all_group_get(self)
        assert groups.status_code == StatusCode.OK

    # def test_get_all_contacts(self):
    #     #Delete all contacts
    #     for x in range(1000):
    #         r = self.__api_get_all_contacts()
    #         assert r.status_code == StatusCode.OK
    #         response = r.json()
    #         json_response = response
    #         for id in json_response['contacts']:
    #             self.__api_delete_contact_request(id['id'])

    def tearDown(self):
        sleep(3)

    if __name__ == "__main__":
        unittest.main()

    def __api_get_list_of_random_hebrew_names(self, number_of_names):
        first_names = ['דנה', 'חנה', 'רינה', 'צילה', 'גילה', 'אבי', 'שלום', 'דני', 'חיים', 'משה']
        last_names = ['כהן', 'לוי', 'בן חיים', 'ציון', 'מזרחי', 'פרץ', 'פרידמן', 'כץ', 'חדד', 'גבאי']
        names = []
        for i in range(number_of_names):
            first = ''.join(random.choice(first_names))
            last = ''.join(random.choice(last_names))
            names.append(f"{first} {last}")
        return names

    def __api_get_list_of_random_emails(self, number_of_emails: int):
        emails = []
        for i in range(number_of_emails):
            email_name = str(uuid.uuid4())
            emails.append(f"{email_name}@comdaUser.co.il")
        return emails

    def __check_hebrew_contacts(self, contacts_list, number_of_contacts):
        for i in range(number_of_contacts):
            contact_details = WesignMethodsApi.contacts_id_get(self, contacts_list[i])
            assert contact_details.status_code == StatusCode.OK
            r = contact_details.json()
            name = r['name']
            assert '?' not in name, "Hebrew name contains ?"
