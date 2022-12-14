import base64
import random
import unittest
import uuid
import warnings
from pathlib import Path
from time import sleep

import openpyxl
import pytest
import requests
import json
from shared import Shared
from status_codes import StatusCode, ResultCode
from all_api_methods import WesignMethodssApi

@pytest.mark.flaky(max_runs=3)
class WesignContactsApi(unittest.TestCase):
    def setUp(self):
        p = Path(__file__).with_name('ContactsSettings.json')
        with open(p) as f:
            self.settings = json.load(f)
        warnings.simplefilter('ignore', ResourceWarning)
        warnings.simplefilter('ignore', DeprecationWarning)
        self.token = Shared.login_request(self)
        ##test

    def test_create_new_contact_by_email_success(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewValidContactWithEmailAndPhone')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactId']
        assert len(json_response) > 0
        r = WesignMethodssApi.contacts_id_delete(self, json_response)
        assert r.status_code == StatusCode.OK

    def test_create_new_contact_by_phone_success(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewValidContactSendingBySms')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactId']
        assert len(json_response) > 0
        r = WesignMethodssApi.contacts_id_delete(self, json_response)
        assert r.status_code == StatusCode.OK

    def test_create_new_contact_by_invalid_email(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewContactWithInvalidEmail')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['Email']
        assert json_response[0] == ResultCode.PLEASE_SPECIFY_VALID_EMAIL

    def test_create_new_contact_by_invalid_phone(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewContactWithInvalidPhone')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['Phone']
        assert json_response[0] == ResultCode.PLEASE_SPECIFY_VALID_PHONE

    def test_create_new_contact_by_empty_name(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewContactWithEmptyName')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['Name']
        assert json_response == ResultCode.INVALID_NAME

    def test_create_new_contact_by_phone_invalid_sending_type(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewValidContactByPhoneInvalidSendingMethod')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['DefaultSendingMethod']
        assert json_response == ResultCode.DEFAULT_SENDING_METHOD

    def test_create_new_contact_by_email_invalid_sending_type(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewValidContactByEmailInvalidSendingMethod')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['DefaultSendingMethod']
        assert json_response == ResultCode.DEFAULT_SENDING_METHOD

    def test_create_new_contact_by_email_invalid_default_sending_method(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewValidContactByEmailInvalidSendingMethod')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['DefaultSendingMethod']
        assert json_response == ResultCode.DEFAULT_SENDING_METHOD

    def test_create_new_contact_with_name_already_exists(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewValidContactNameExists')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['error']
        assert json_response[0] == ResultCode.CONTACT_WITH_SAME_MEANS_ALREADY_EXISTS

    def test_create_new_contact_with_email_already_exists(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewValidContactEmailExists')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['error']
        assert json_response[0] == ResultCode.CONTACT_WITH_SAME_MEANS_ALREADY_EXISTS

    def test_create_new_contact_with_phone_already_exists(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewValidContactPhoneExists')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['error']
        assert json_response[0] == ResultCode.CONTACT_WITH_SAME_MEANS_ALREADY_EXISTS

    def test_create_new_contact_from_bulk_xlsx_file_success(self):
        r = WesignMethodssApi.contacts_bulk_post_json_file(self, 'CreateNewContactFromBulkCsv')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactsId']
        assert len(json_response[0]) == 36
        r = WesignMethodssApi.contacts_id_delete(self, json_response[0])
        assert r.status_code == StatusCode.OK

    def test_create_new_contacts_from_bulk_xlsx_sending_method_sms_success(self):
        r = WesignMethodssApi.contacts_bulk_post_json_file(self, 'CreateNewContactsFromBulkCsvSendingMehtodSms')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactsId']
        assert len(json_response[0]) == 36
        for contacts in json_response:
            WesignMethodssApi.contacts_id_delete(self, contacts)

    def test_create_new_contacts_from_bulk_xlsx_sending_method_email_success(self):
        r = WesignMethodssApi.contacts_bulk_post_json_file(self, 'CreateNewContactsFromBulkCsvSendingMehtodEmail')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactsId']
        assert len(json_response[0]) == 36
        for contacts in json_response:
            WesignMethodssApi.contacts_id_delete(self, contacts)

    def test_create_new_contacts_from_bulk_xlsx_sending_method_email_and_sms_success(self):
        r = WesignMethodssApi.contacts_bulk_post_json_file(self, 'CreateNewContactsFromBulkCsvSendingMehtodEmailAndSms')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactsId']
        assert len(json_response[0]) == 36
        for contacts in json_response:
            WesignMethodssApi.contacts_id_delete(self, contacts)

    def test_create_new_contact_from_bulk_xlsx_valid_name_and_email_invalid_phone(self):
        r = WesignMethodssApi.contacts_bulk_post_json_file(self, 'CreateNewContactsFromBulkCsvInvalidNameAndEmailInvalidPhone')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['error']
        assert json_response[0] == ResultCode.INVALID_PHONE

    def test_create_new_contact_from_bulk_xlsx_empty_csv(self):
        r = WesignMethodssApi.contacts_bulk_post_json_file(self, 'CreateNewContactsFromBulkCsvEmpty')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['error']
        assert json_response[0] == ResultCode.INVALID_FORMAT

    def test_create_new_contact_from_bulk_csv_empty_fullname_valid_phone(self):
        r = WesignMethodssApi.contacts_bulk_post_json_file(self, 'CreateNewContactsFromBulkCsvEmptyFullnameValidPhone')
        assert r.status_code == StatusCode.BAD_REQUEST
        response = r.json()
        json_response = response['errors']['error']
        assert json_response[0] == ResultCode.NAME_IS_MISSING

    def test_update_contact_mail(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewValidContactWithEmailAndPhone')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactId']
        r = WesignMethodssApi.contacts_id_put(self, json_response, 'UpdateContactMail')
        assert r.status_code == StatusCode.OK
        WesignMethodssApi.contacts_id_delete(self, json_response)
        assert r.status_code == StatusCode.OK

    def test_update_contact_phone(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewValidContactWithEmailAndPhone')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactId']
        r = WesignMethodssApi.contacts_id_put(self, json_response, 'UpdateContactPhone')
        assert r.status_code == StatusCode.OK
        WesignMethodssApi.contacts_id_delete(self, json_response)
        assert r.status_code == StatusCode.OK

    def test_update_contact_seal(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewValidContactWithEmailAndPhone')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactId']
        r = WesignMethodssApi.contacts_id_put(self, json_response, 'UpdateContactSeal')
        assert r.status_code == StatusCode.OK
        WesignMethodssApi.contacts_id_delete(self, json_response)
        assert r.status_code == StatusCode.OK

    def test_update_contact_invalid_email(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewValidContactWithEmailAndPhone')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactId']
        r = WesignMethodssApi.contacts_id_put(self, json_response, 'UpdateContactInvalidPhone')
        assert r.status_code == StatusCode.BAD_REQUEST
        response_error = r.json()
        json_response_error = response_error['errors']['Phone']
        assert json_response_error[0] == ResultCode.PLEASE_SPECIFY_VALID_PHONE
        r = WesignMethodssApi.contacts_id_delete(self, json_response)
        assert r.status_code == StatusCode.OK

    @pytest.mark.skip('Not implement in 3.0.8 version')
    def test_update_contact_invalid_name(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewValidContactWithEmailAndPhone')
        assert r.status_code == StatusCode.OK
        response = r.json()
        json_response = response['contactId']
        r = WesignMethodssApi.contacts_id_put(self, json_response, 'UpdateContactInvalidName')
        assert r.status_code == StatusCode.BAD_REQUEST
        response_error = r.json()
        json_response_error = response_error['errors']['Name']
        assert json_response_error[0] == ResultCode.NAME_SHOULD_CONTAIN_ONLY_CHARACTERS
        r = WesignMethodssApi.contacts_id_delete(self, json_response)
        assert r.status_code == StatusCode.OK

    #Bug number = WES-1047
    def test_create_new_contact_with_global_number_when_provider_is_goldman(self):
        r = WesignMethodssApi.contacts_post_json_file(self, 'CreateNewContactWithGlobalPhoneWhenProviderIsGoldman')
        assert r.status_code == StatusCode.BAD_REQUEST
        response_error = r.json()
        json_response_error = response_error['errors']['error']
        assert json_response_error[0] == ResultCode.SMS_PROVIDER_ERROR

    def test_delete_multi_contacts_by_id_success(self):
        a = []
        for x in range(4):
            contact = uuid.uuid4().hex
            self.contact_name = contact
            d = {
                  "name": self.contact_name,
                  "email": self.contact_name + "@ApiUser.com",
                  "defaultSendingMethod": 2
                }
            r = WesignMethodssApi.contacts_post(self, d)
            assert r.status_code == StatusCode.OK
            response = r.json()
            json_response = response['contactId']
            assert len(json_response) > 0
            a.append(response['contactId'])
        c = {
              "ids": [
                    a[0], a[1], a[2], a[3]
              ]
            }
        delete = WesignMethodssApi.contacts_delete_batch_put(self, c)
        assert delete.status_code == StatusCode.OK

    def test_upload_hebrew_contacts_from_xlsx_file_and_check_names(self):
        number_of_contacts = 10
        list_of_names = self.__api_get_list_of_random_hebrew_names(number_of_contacts)
        list_of_emails = self.__api_get_list_of_random_emails(number_of_contacts)
        book = openpyxl.load_workbook(self.settings['openXlsxFileForUploadContacts'])
        sheet = book.active
        for i in range(number_of_contacts):
            sheet.cell(row=i+2, column=1).value = list_of_names[i]  # FullName
            sheet.cell(row=i + 2, column=2).value = list_of_emails[i]  # Email
            sheet.cell(row=i + 2, column=4).value = 2
        book.save(self.settings['openXlsxFileForUploadContacts'])
        data = open(self.settings['openXlsxFileForUploadContacts'], 'rb').read()
        base64_encoded = base64.b64encode(data).decode('UTF-8')
        base64_string = '"base64File": "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,' + base64_encoded
        contacts_id = WesignMethodssApi.contacts_bulk_post_xlsx_file(self, base64_string)
        assert contacts_id.status_code == StatusCode.OK
        response = contacts_id.json()
        contacts_list = response['contactsId']
        self.__check_hebrew_contacts(contacts_list, number_of_contacts)
        sheet.delete_rows(2, number_of_contacts)  # Delete contacts from xlsx file
        book.save(self.settings['openXlsxFileForUploadContacts'])
        for contact_id in contacts_list:
            WesignMethodssApi.contacts_id_delete(self, contact_id)

    # def test_get_all_contacts(self):
    #     #Delete all contacts
    #     for x in range(1000):
    #         r = self.__api_get_all_contacts()
    #         assert r.status_code == StatusCode.OK
    #         response = r.json()
    #         json_response = response
    #         for id in json_response['contacts']:
    #             self.__api_delete_contact_request(id['id'])

    def tearDown(self):
        sleep(1)

    if __name__ == "__main__":
        unittest.main()

    def __api_get_list_of_random_hebrew_names(self, number_of_names):
        first_names = ['דנה', 'חנה', 'רינה', 'צילה', 'גילה', 'אבי', 'שלום', 'דני', 'חיים', 'משה']
        last_names = ['כהן', 'לוי', 'בן חיים', 'ציון', 'מזרחי', 'פרץ', 'פרידמן', 'כץ', 'חדד', 'גבאי']
        names = []
        for i in range(number_of_names):
            first = ''.join(random.choice(first_names))
            last = ''.join(random.choice(last_names))
            names.append(f"{first} {last}")
        return names

    def __api_get_list_of_random_emails(self, number_of_emails: int):
        emails = []
        for i in range(number_of_emails):
            email_name = str(uuid.uuid4())
            emails.append(f"{email_name}@comdaUser.co.il")
        return emails

    def __check_hebrew_contacts(self, contacts_list, number_of_contacts):
        for i in range(number_of_contacts):
            contact_details = WesignMethodssApi.contacts_id_get(self, contacts_list[i])
            assert contact_details.status_code == StatusCode.OK
            r = contact_details.json()
            name = r['name']
            assert '?' not in name, "Hebrew name contains ?"
